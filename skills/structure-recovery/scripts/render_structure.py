#!/usr/bin/env python3
"""render_structure.py — deterministic structure-catalog renderer (HTML + CSV + inferred DDL).

structure-recovery WP-7. Consumes a ``structure-index.v1`` catalog (the
cross-file accumulated structure index — entities[] with COMPUTED byte layout,
resolved relationships[], gaps[]) and produces three stdlib-only outputs:

- structure.html  — one schema table per entity + a relationship DAG via
                    Cytoscape, confidence colour-coded. Deterministic under
                    SOURCE_DATE_EPOCH, XSS-safe (html.escape + JS textContent),
                    vendored-Cytoscape air-gap fallback. Mirrors the shape of
                    lineage-extract-static/scripts/render_report.py (:363-486)
                    but is STDLIB-ONLY (no Jinja2) because the structure
                    catalog's primary surface is the per-entity tables, not the
                    graph — so the HTML ALWAYS renders, degrading the graph
                    section to a banner when Cytoscape cannot load.
- fields.csv / relationships.csv — flat exports.
- structure.ddl.sql — INFERRED DDL, advisory-only. Carries a mandatory
                    human-review header; LIVE constraints are emitted ONLY at
                    grounded confidence; inferred/speculative relationships are
                    emitted as COMMENTED ``-- INFERRED FK:`` lines, NEVER as live
                    constraints, NEVER executed.

Excel + wiki emitters are WP-8 (they compose ms-office-excel-python / wiki and
call the render_csv / render_ddl seams here). OpenLineage SchemaDatasetFacet is
WP-9 (extends merge_into_ol.py and calls into this catalog).

SAFETY (design §3.3 / §5 / Codex Finding 1):
- A field whose offset is RANGED (post-ODO position, SYNCHRONIZED slack) is
  rendered as ``min..max`` (or ``?`` when unknown) — NEVER a single confident
  value.
- A variable_length entity is bannered ``record length varies``; its DDL omits a
  fixed record-length assertion.
- Inferred/speculative FKs are advisory; they never become a live constraint and
  never feed a gate.

CLI:
    render_structure.py <catalog.json> --output-dir <dir>
                        [--no-vendor] [--project-name NAME]
                        [--source-date-epoch UNIX_TIMESTAMP]
"""

from __future__ import annotations

import argparse
import datetime as _dt
import html
import json
import os
import socket
import sys
import tempfile
from pathlib import Path
from typing import Any, Iterable, Optional, Sequence

EXTRACTOR_ID = "structure-recovery"
EXTRACTOR_VERSION = "1.0.0"

# Air-gap-safe vendor lookup (shared vendor dir with the lineage renderer).
VENDOR_DIR = Path.home() / ".claude" / "skills" / "visual-companion" / "templates" / "vendor"
CYTOSCAPE_VENDOR = VENDOR_DIR / "cytoscape.min.js"

CDN_CYTOSCAPE = "https://unpkg.com/cytoscape@3.28.1/dist/cytoscape.min.js"

# Confidence -> colour (used by both the HTML legend and per-cell badges).
_CONF_COLORS = {
    "grounded": "#1a7f37",      # green
    "inferred": "#9a6700",      # amber
    "speculative": "#cf222e",   # red
}
_CONF_ORDER = {"grounded": 0, "inferred": 1, "speculative": 2}


# ---------------------------------------------------------------------------
# Atomic write (reused shape from render_report.atomic_write_text)
# ---------------------------------------------------------------------------

def atomic_write_text(path: Path, content: str) -> None:
    """Write text atomically via .tmp.<pid> + os.replace + fsync."""
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_path = tempfile.mkstemp(
        prefix=path.name + ".tmp.",
        suffix=f".{os.getpid()}",
        dir=str(path.parent),
    )
    try:
        with os.fdopen(fd, "w", encoding="utf-8", newline="\n") as f:
            f.write(content)
            f.flush()
            os.fsync(f.fileno())
        os.replace(tmp_path, path)
    except Exception:
        try:
            os.unlink(tmp_path)
        except FileNotFoundError:
            pass
        raise


# ---------------------------------------------------------------------------
# Determinism + air-gap helpers
# ---------------------------------------------------------------------------

def _resolve_source_date_epoch(explicit: Optional[int]) -> Optional[int]:
    """Resolve the SOURCE_DATE_EPOCH (explicit arg wins, then env)."""
    if explicit is not None:
        return explicit
    raw = os.environ.get("SOURCE_DATE_EPOCH")
    if raw:
        try:
            return int(raw)
        except ValueError:
            return None
    return None


def _generated_at(source_date_epoch: Optional[int]) -> str:
    """A deterministic ISO-8601 UTC stamp when SOURCE_DATE_EPOCH is set,
    else wall-clock. Determinism (Reproducible Builds convention)."""
    if source_date_epoch is not None:
        ts = _dt.datetime.fromtimestamp(source_date_epoch, tz=_dt.timezone.utc)
        return ts.strftime("%Y-%m-%dT%H:%M:%SZ")
    return _dt.datetime.now(tz=_dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _check_cdn_reachable(host: str = "unpkg.com", port: int = 443, timeout: float = 2.0) -> bool:
    """DNS + TCP probe. Hint only — the real fetch is in the browser.

    Honours HOST_NETWORK_AVAILABLE for hermetic tests / forced air-gap.
    """
    flag = os.environ.get("HOST_NETWORK_AVAILABLE", "").lower()
    if flag in ("false", "0", "no"):
        return False
    if flag in ("true", "1", "yes"):
        return True
    try:
        socket.setdefaulttimeout(timeout)
        with socket.create_connection((host, port), timeout=timeout):
            return True
    except (socket.timeout, socket.gaierror, OSError):
        return False


def _resolve_cytoscape_src(output_dir: Path, no_vendor: bool) -> tuple[Optional[str], bool, bool]:
    """Resolve the cytoscape <script src>.

    Returns (src, cdn_fallback, graph_available):
      - vendored present -> (relpath, False, True)
      - no vendor, CDN reachable, not --no-vendor -> (CDN_url, True, True)
      - otherwise -> (None, False, False)  # graph degraded to a banner
    """
    if CYTOSCAPE_VENDOR.exists():
        try:
            return os.path.relpath(CYTOSCAPE_VENDOR, output_dir), False, True
        except ValueError:
            return str(CYTOSCAPE_VENDOR), False, True
    if no_vendor:
        return None, False, False
    if _check_cdn_reachable():
        return CDN_CYTOSCAPE, True, True
    return None, False, False


# ---------------------------------------------------------------------------
# Catalog access (tolerant of the structure-index.v1 shape and of None fields)
# ---------------------------------------------------------------------------

def _entities(catalog: dict) -> list[dict]:
    return list(catalog.get("entities") or [])


def _relationships(catalog: dict) -> list[dict]:
    return list(catalog.get("relationships") or [])


def _catalog_gaps(catalog: dict) -> list[dict]:
    return list(catalog.get("gaps") or [])


def _entity_sort_key(e: dict) -> tuple:
    return (str(e.get("object_kind", "")), str(e.get("qualified_name", "")))


def _sorted_entities(catalog: dict) -> list[dict]:
    return sorted(_entities(catalog), key=_entity_sort_key)


def _sorted_fields(entity: dict) -> list[dict]:
    """Fields in declaration order (ordinal, then name) — deterministic."""
    fields = list(entity.get("fields") or [])
    return sorted(fields, key=lambda f: (f.get("ordinal", 0), str(f.get("name", ""))))


def _rel_sort_key(r: dict) -> tuple:
    return (
        str(r.get("kind", "")),
        str(r.get("from_object", "")),
        str(r.get("from_field", "")),
        str(r.get("to_object") or ""),
        str(r.get("to_field") or ""),
    )


def _sorted_relationships(catalog: dict) -> list[dict]:
    return sorted(_relationships(catalog), key=_rel_sort_key)


def _offset_display(field: dict) -> str:
    """Render a field's byte offset HONESTLY.

    - ranged / no single authoritative offset -> 'min..max' (or '?')
    - single authoritative offset -> str(offset)
    - unknown -> ''
    (Codex Finding 1 — NEVER a single confident value for a ranged offset.)
    """
    if field.get("ranged"):
        lo = field.get("byte_offset_min")
        hi = field.get("byte_offset_max")
        lo_s = str(lo) if lo is not None else "?"
        hi_s = str(hi) if hi is not None else "?"
        return f"{lo_s}..{hi_s}"
    off = field.get("byte_offset")
    if off is not None:
        return str(off)
    # not ranged, no single offset: fall back to a min..max if present, else blank
    lo = field.get("byte_offset_min")
    hi = field.get("byte_offset_max")
    if lo is not None or hi is not None:
        return f"{(lo if lo is not None else '?')}..{(hi if hi is not None else '?')}"
    return ""


def _length_display(field: dict) -> str:
    if field.get("variable_length"):
        ln = field.get("length")
        return f"{ln} (var)" if ln is not None else "var"
    ln = field.get("length")
    return str(ln) if ln is not None else ""


def _field_type_display(field: dict) -> str:
    """Best human-facing type label: declared_type, else pic_clause, else
    normalized_type, else ''."""
    for key in ("declared_type", "pic_clause", "normalized_type"):
        val = field.get(key)
        if val:
            return str(val)
    return ""


# ---------------------------------------------------------------------------
# Cytoscape elements (relationship DAG): entities are nodes, fk/join are edges
# ---------------------------------------------------------------------------

def _safe_node_id(s: str) -> str:
    out = "".join(c if c.isalnum() else "_" for c in s)
    if len(out) > 60:
        import hashlib
        out = out[:50] + "_" + hashlib.sha1(s.encode()).hexdigest()[:8]
    return out


def build_cytoscape_elements(catalog: dict) -> list[dict]:
    """Entities -> nodes; fk/join relationships with a resolved to_object -> edges.

    pk/unique relationships are entity-internal (no edge). An unresolved fk
    (to_object is None) gets no edge. Deterministic ordering throughout.
    """
    elements: list[dict] = []
    entity_ids: dict[str, str] = {}
    for e in _sorted_entities(catalog):
        qn = str(e.get("qualified_name", ""))
        node_id = f"e_{_safe_node_id(qn)}"
        entity_ids[qn] = node_id
        elements.append({
            "data": {
                "id": node_id,
                "label": qn,
                "object_kind": str(e.get("object_kind", "")),
                "confidence": str(e.get("confidence", "")),
            }
        })
    edge_i = 0
    for r in _sorted_relationships(catalog):
        if r.get("kind") not in ("fk", "join"):
            continue
        src_qn = str(r.get("from_object", ""))
        dst_qn = r.get("to_object")
        if not dst_qn:
            continue
        src_id = entity_ids.get(src_qn)
        dst_id = entity_ids.get(str(dst_qn))
        if not src_id or not dst_id:
            # endpoint not in the catalog — don't fabricate a node
            continue
        elements.append({
            "data": {
                "id": f"rel_{edge_i}",
                "source": src_id,
                "target": dst_id,
                "kind": str(r.get("kind", "")),
                "confidence": str(r.get("confidence", "")),
                "label": f"{r.get('kind','')}:{r.get('from_field','')}",
            }
        })
        edge_i += 1
    return elements


def _embed_json_xss_safe(obj: Any) -> str:
    """Serialize to JSON and escape the characters that could break out of a
    <script> context. Mirrors render_report.py:455-462. HARD-RULE 7."""
    raw = json.dumps(obj, ensure_ascii=False, sort_keys=True)
    return (
        raw.replace("<", "\\u003c")
           .replace(">", "\\u003e")
           .replace("&", "\\u0026")
           .replace(" ", "\\u2028")
           .replace(" ", "\\u2029")
    )


# ---------------------------------------------------------------------------
# Markdown escaping (for the wiki emitter — keep tables/inline structurally sound)
# ---------------------------------------------------------------------------

def _md_inline(value: Any) -> str:
    """Neutralise characters that would break markdown INLINE text flow.
    Collapses CR/LF to a space (a stray newline would split a list item) and
    leaves backticks intact (we wrap code spans ourselves around known-safe
    short tokens, never around free text)."""
    s = "" if value is None else str(value)
    return s.replace("\r", " ").replace("\n", " ")


def _md_cell(value: Any) -> str:
    """Neutralise a value for a GFM table cell. A literal ``|`` would inject an
    extra column and a newline would terminate the row — escape the pipe and
    flatten CR/LF. Returns the cell-safe string (no surrounding pipes)."""
    return _md_inline(value).replace("|", "\\|")


# ---------------------------------------------------------------------------
# HTML
# ---------------------------------------------------------------------------

_CSS = """
:root { color-scheme: light dark; }
* { box-sizing: border-box; }
body { font: 14px/1.5 -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
       margin: 0; padding: 1.5rem; max-width: 1200px; margin: 0 auto; }
h1 { font-size: 1.5rem; margin: 0 0 .25rem; }
h2 { font-size: 1.15rem; margin: 2rem 0 .5rem; border-bottom: 1px solid #d0d7de; padding-bottom: .25rem; }
.meta { color: #57606a; font-size: 12px; margin-bottom: 1rem; }
.badge { display: inline-block; padding: 1px 6px; border-radius: 8px; font-size: 11px;
         color: #fff; font-weight: 600; }
table { border-collapse: collapse; width: 100%; margin: .5rem 0 1rem; font-size: 13px; }
th, td { border: 1px solid #d0d7de; padding: 4px 8px; text-align: left; vertical-align: top; }
th { background: #f6f8fa; position: sticky; top: 0; }
td.num { text-align: right; font-variant-numeric: tabular-nums; }
.entity { margin-bottom: .5rem; }
.kindtag { color: #57606a; font-weight: 400; font-size: 12px; }
.gaps { color: #9a6700; }
.gaps li { margin: 2px 0; }
.cite { color: #57606a; font-size: 11px; }
.legend span { margin-right: 1rem; }
#graph { width: 100%; height: 480px; border: 1px solid #d0d7de; border-radius: 6px; }
.banner { background: #fff8c5; border: 1px solid #d4a72c; border-radius: 6px;
          padding: 8px 12px; margin: .5rem 0; font-size: 13px; }
.ranged { font-style: italic; color: #9a6700; }
""".strip()


def _confidence_badge_html(conf: str) -> str:
    conf = conf or "speculative"
    color = _CONF_COLORS.get(conf, "#57606a")
    return f'<span class="badge" style="background:{html.escape(color)}">{html.escape(conf)}</span>'


def _legend_html() -> str:
    parts = ['<p class="legend">']
    for conf in ("grounded", "inferred", "speculative"):
        parts.append(
            f'<span>{_confidence_badge_html(conf)} {html.escape(conf)}</span>'
        )
    parts.append("</p>")
    return "".join(parts)


def _gaps_list_html(gaps: Iterable[dict]) -> str:
    items = []
    for g in sorted(gaps, key=lambda x: (str(x.get("kind", "")), str(x.get("file_path") or ""),
                                         x.get("line") or 0, str(x.get("description", "")))):
        kind = html.escape(str(g.get("kind", "")))
        desc = html.escape(str(g.get("description", "")))
        loc = ""
        fp = g.get("file_path")
        ln = g.get("line")
        if fp:
            loc = f' <span class="cite">({html.escape(str(fp))}{":" + html.escape(str(ln)) if ln else ""})</span>'
        items.append(f"<li><code>{kind}</code> — {desc}{loc}</li>")
    if not items:
        return ""
    return '<ul class="gaps">' + "".join(items) + "</ul>"


def _entity_table_html(entity: dict) -> str:
    qn = html.escape(str(entity.get("qualified_name", "")))
    kind = html.escape(str(entity.get("object_kind", "")))
    conf = str(entity.get("confidence", ""))
    ev = entity.get("evidence") or {}
    cite = ""
    if ev.get("file_path"):
        cite = (f' <span class="cite">{html.escape(str(ev.get("file_path")))}'
                f':{html.escape(str(ev.get("line", "")))}</span>')

    # record length banner
    rl_banner = ""
    if entity.get("variable_length"):
        lo = entity.get("record_length_min")
        hi = entity.get("record_length_max")
        rng = f"{lo if lo is not None else '?'}..{hi if hi is not None else '?'}"
        rl_banner = f'<div class="banner">Record length varies (ODO): {html.escape(rng)} bytes — no single authoritative length.</div>'
    elif entity.get("record_length") is not None:
        rl_banner = f'<div class="meta">Record length: {html.escape(str(entity.get("record_length")))} bytes</div>'

    head = (f'<div class="entity"><strong>{qn}</strong> '
            f'<span class="kindtag">[{kind}]</span> {_confidence_badge_html(conf)}{cite}</div>')

    rows = []
    has_cobol = entity.get("object_kind") == "cobol_record"
    for f in _sorted_fields(entity):
        name = html.escape(str(f.get("name", "")))
        typ = html.escape(_field_type_display(f))
        off = _offset_display(f)
        off_cls = "num ranged" if f.get("ranged") else "num"
        off_html = html.escape(off)
        length = html.escape(_length_display(f))
        fconf = str(f.get("confidence", ""))
        ev_kind = html.escape(str(f.get("evidence_kind", "")))
        enf = html.escape(str(f.get("enforcement", "")))
        level = f.get("level")
        level_html = html.escape(str(level)) if level is not None else ""
        nullable = f.get("nullable")
        null_html = "" if nullable is None else ("YES" if nullable else "NO")
        if has_cobol:
            rows.append(
                f"<tr><td>{name}</td><td>{level_html}</td><td>{typ}</td>"
                f'<td class="{off_cls}">{off_html}</td><td class="num">{length}</td>'
                f"<td>{_confidence_badge_html(fconf)}</td><td>{ev_kind}</td></tr>"
            )
        else:
            rows.append(
                f"<tr><td>{name}</td><td>{typ}</td><td>{html.escape(str(null_html))}</td>"
                f"<td>{_confidence_badge_html(fconf)}</td><td>{ev_kind}</td><td>{enf}</td></tr>"
            )

    if has_cobol:
        header = ("<tr><th>field</th><th>level</th><th>type</th><th>offset</th>"
                  "<th>length</th><th>confidence</th><th>evidence</th></tr>")
    else:
        header = ("<tr><th>column</th><th>type</th><th>nullable</th>"
                  "<th>confidence</th><th>evidence</th><th>enforcement</th></tr>")

    table = ""
    if rows:
        table = f"<table><thead>{header}</thead><tbody>{''.join(rows)}</tbody></table>"
    else:
        table = '<div class="meta">No fields (schema not derivable — see gaps).</div>'

    entity_gaps = _gaps_list_html(entity.get("gaps") or [])
    return head + rl_banner + table + entity_gaps


_GRAPH_JS = r"""
(function () {
  var data = window.__STRUCTURE_ELEMENTS__ || [];
  if (typeof cytoscape === "undefined") {
    var el = document.getElementById("graph");
    if (el) { el.textContent = "Relationship graph unavailable (Cytoscape did not load)."; }
    return;
  }
  var confColor = {grounded: "#1a7f37", inferred: "#9a6700", speculative: "#cf222e"};
  cytoscape({
    container: document.getElementById("graph"),
    elements: data,
    style: [
      { selector: "node", style: {
          "label": "data(label)", "font-size": "10px", "text-wrap": "wrap",
          "text-max-width": "120px", "background-color": "#0969da",
          "color": "#24292f", "text-valign": "bottom" } },
      { selector: "edge", style: {
          "width": 2, "curve-style": "bezier",
          "target-arrow-shape": "triangle",
          "line-color": function (e) { return confColor[e.data("confidence")] || "#57606a"; },
          "target-arrow-color": function (e) { return confColor[e.data("confidence")] || "#57606a"; },
          "label": "data(label)", "font-size": "9px" } }
    ],
    layout: { name: "cose", animate: false }
  });
})();
""".strip()


def render_html(
    catalog: dict,
    output_dir: Path,
    *,
    no_vendor: bool = False,
    project_name: str = "project",
    source_date_epoch: Optional[int] = None,
) -> Path:
    """Render structure.html (ALWAYS — the per-entity tables are the primary
    surface; the relationship graph degrades to a banner if Cytoscape cannot
    load). Returns the output path."""
    sde = _resolve_source_date_epoch(source_date_epoch)
    output_dir.mkdir(parents=True, exist_ok=True)

    cyto_src, cdn_fallback, graph_available = _resolve_cytoscape_src(output_dir, no_vendor)
    entities = _sorted_entities(catalog)
    relationships = _sorted_relationships(catalog)
    elements = build_cytoscape_elements(catalog)

    gw = catalog.get("generated_with") or {}
    partial = bool(gw.get("partial"))
    infer_rel = bool(gw.get("infer_relationships"))

    parts: list[str] = []
    parts.append("<!DOCTYPE html>")
    parts.append('<html lang="en"><head><meta charset="utf-8">')
    parts.append('<meta name="viewport" content="width=device-width, initial-scale=1">')
    parts.append(f"<title>Structure catalog — {html.escape(project_name)}</title>")
    parts.append(f"<style>{_CSS}</style></head><body>")

    parts.append(f"<h1>Structure catalog — {html.escape(project_name)}</h1>")
    parts.append(
        f'<p class="meta">Generated by {html.escape(EXTRACTOR_ID)} '
        f"v{html.escape(EXTRACTOR_VERSION)} at {html.escape(_generated_at(sde))} · "
        f"{len(entities)} entit{'y' if len(entities) == 1 else 'ies'} · "
        f"{len(relationships)} relationship{'' if len(relationships) == 1 else 's'} · "
        f"infer_relationships={'on' if infer_rel else 'off'}</p>"
    )
    if partial:
        parts.append('<div class="banner">PARTIAL run (wall-clock cap reached) — this catalog is incomplete.</div>')

    parts.append('<div class="banner">Inferred structure — characterization aid; offsets/types/relationships are static-analysis derived and require human review. Inferred and speculative relationships are advisory only.</div>')
    parts.append(_legend_html())

    # Relationship graph
    parts.append("<h2>Relationships</h2>")
    if cdn_fallback:
        parts.append('<div class="banner">Cytoscape loaded from CDN (vendored copy not found). The graph requires network access in the browser.</div>')
    if not graph_available:
        parts.append('<div class="banner">Relationship graph not rendered (no vendored Cytoscape and no network). The relationship table below is authoritative.</div>')
    parts.append('<div id="graph"></div>')

    # Relationship table (always — authoritative even when the graph can't draw)
    if relationships:
        rrows = []
        for r in relationships:
            rrows.append(
                "<tr>"
                f"<td>{html.escape(str(r.get('kind', '')))}</td>"
                f"<td>{html.escape(str(r.get('from_object', '')))}</td>"
                f"<td>{html.escape(str(r.get('from_field', '')))}</td>"
                f"<td>{html.escape(str(r.get('to_object') or ''))}</td>"
                f"<td>{html.escape(str(r.get('to_field') or ''))}</td>"
                f"<td>{_confidence_badge_html(str(r.get('confidence', '')))}</td>"
                f"<td>{html.escape(str(r.get('enforcement', '')))}</td>"
                "</tr>"
            )
        parts.append(
            "<table><thead><tr><th>kind</th><th>from object</th><th>from field</th>"
            "<th>to object</th><th>to field</th><th>confidence</th><th>enforcement</th>"
            f"</tr></thead><tbody>{''.join(rrows)}</tbody></table>"
        )
    else:
        parts.append('<p class="meta">No relationships resolved.</p>')

    # Entities
    parts.append("<h2>Entities</h2>")
    if entities:
        for e in entities:
            parts.append(_entity_table_html(e))
    else:
        parts.append('<p class="meta">No entities in catalog.</p>')

    # Catalog-level gaps
    cat_gaps = _catalog_gaps(catalog)
    if cat_gaps:
        parts.append("<h2>Catalog gaps</h2>")
        parts.append(_gaps_list_html(cat_gaps))

    # Embed elements + scripts. textContent-only JS (never innerHTML). HARD-RULE 7.
    parts.append(
        f'<script>window.__STRUCTURE_ELEMENTS__ = {_embed_json_xss_safe(elements)};</script>'
    )
    if cyto_src:
        parts.append(f'<script src="{html.escape(cyto_src)}"></script>')
    parts.append(f"<script>{_GRAPH_JS}</script>")
    parts.append("</body></html>")
    parts.append("")  # trailing newline (keep_trailing_newline parity)

    out_path = output_dir / "structure.html"
    atomic_write_text(out_path, "\n".join(parts))
    return out_path


# ---------------------------------------------------------------------------
# CSV (stdlib; deterministic; RFC-4180-ish quoting via csv module)
# ---------------------------------------------------------------------------

def _csv_escape(value: Any) -> str:
    """Quote a value for CSV if it contains , " \r or \n; double internal quotes.
    Also neutralises a leading formula trigger (=,+,-,@) by prefixing a single
    quote — CSVs opened in a spreadsheet are an injection vector too (CWE-1236).
    """
    s = "" if value is None else str(value)
    if s and s[0] in ("=", "+", "-", "@"):
        s = "'" + s
    if any(c in s for c in (",", '"', "\r", "\n")):
        s = '"' + s.replace('"', '""') + '"'
    return s


def _csv_row(values: Sequence[Any]) -> str:
    return ",".join(_csv_escape(v) for v in values)


def render_csv(catalog: dict, output_dir: Path) -> dict[str, Path]:
    """Write fields.csv and relationships.csv. Returns {name: path}.

    fields.csv columns: entity, field, type, offset, length, confidence, evidence_kind
    relationships.csv columns: kind, from_object, from_field, to_object, to_field, confidence, enforcement
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    field_lines = ["entity,field,type,offset,length,confidence,evidence_kind"]
    for e in _sorted_entities(catalog):
        qn = e.get("qualified_name", "")
        for f in _sorted_fields(e):
            field_lines.append(_csv_row([
                qn,
                f.get("name", ""),
                _field_type_display(f),
                _offset_display(f),
                _length_display(f),
                f.get("confidence", ""),
                f.get("evidence_kind", ""),
            ]))
    fields_path = output_dir / "fields.csv"
    atomic_write_text(fields_path, "\n".join(field_lines) + "\n")

    rel_lines = ["kind,from_object,from_field,to_object,to_field,confidence,enforcement"]
    for r in _sorted_relationships(catalog):
        rel_lines.append(_csv_row([
            r.get("kind", ""),
            r.get("from_object", ""),
            r.get("from_field", ""),
            r.get("to_object") or "",
            r.get("to_field") or "",
            r.get("confidence", ""),
            r.get("enforcement", ""),
        ]))
    rel_path = output_dir / "relationships.csv"
    atomic_write_text(rel_path, "\n".join(rel_lines) + "\n")

    return {"fields.csv": fields_path, "relationships.csv": rel_path}


# ---------------------------------------------------------------------------
# Inferred DDL (.sql) — ADVISORY ONLY
# ---------------------------------------------------------------------------

_DDL_HEADER = (
    "-- INFERRED DDL — characterization aid, requires human review\n"
    "-- Generated by structure-recovery (static analysis). NOT a migration.\n"
    "-- LIVE constraints below are emitted only at GROUNDED confidence.\n"
    "-- Inferred/speculative relationships appear as commented `-- INFERRED FK:` lines and are NEVER executed.\n"
    "-- DO NOT run this against a database without review.\n"
)

# SQL identifier safe-ish: keep [A-Za-z0-9_.] and the original; quote otherwise.
def _ddl_ident(name: str) -> str:
    s = str(name)
    safe = all(c.isalnum() or c in ("_", ".") for c in s) and bool(s)
    if safe:
        return s
    # double-quote and escape embedded quotes (standard SQL delimited identifier)
    return '"' + s.replace('"', '""') + '"'


def _ddl_comment_safe(text: str) -> str:
    """A comment must not contain a newline (would break out of the -- line)."""
    return str(text).replace("\r", " ").replace("\n", " ")


def _is_grounded_declared_fk(r: dict) -> bool:
    """A relationship that earns a LIVE FOREIGN KEY: a declared, grounded,
    fully-resolved fk on a relational table. Everything else is advisory."""
    return bool(
        r.get("kind") == "fk"
        and r.get("confidence") == "grounded"
        and r.get("enforcement") == "declared"
        and r.get("to_object")
        and r.get("to_field")
    )


def _inferred_fk_comment(qn: str, r: dict) -> str:
    """One commented `-- INFERRED FK:` advisory line (never executed)."""
    tgt = f"{r.get('to_object') or '?'}.{r.get('to_field') or '?'}"
    return (
        f"-- INFERRED FK: {_ddl_comment_safe(qn)}.{_ddl_comment_safe(r.get('from_field',''))} "
        f"-> {_ddl_comment_safe(tgt)} "
        f"[{r.get('kind','')}, {r.get('confidence','')}] -- advisory, never executed"
    )


def render_ddl(catalog: dict, output_dir: Path) -> Path:
    """Write structure.ddl.sql. Advisory-only inferred DDL.

    - CREATE TABLE per table/view entity with grounded fields.
    - Live PK/UNIQUE/FK constraints ONLY when the relationship confidence is
      grounded AND enforcement is declared.
    - Inferred/speculative relationships -> commented `-- INFERRED FK:` lines.
    - COBOL records and flat-file layouts are rendered as commented field
      manifests (no live DDL — they are not relational tables).
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    out: list[str] = [_DDL_HEADER]

    entities = _sorted_entities(catalog)
    relationships = _sorted_relationships(catalog)

    # Index relationships by owning object for inlining.
    rels_by_obj: dict[str, list[dict]] = {}
    for r in relationships:
        rels_by_obj.setdefault(str(r.get("from_object", "")), []).append(r)

    for e in entities:
        kind = e.get("object_kind")
        qn = str(e.get("qualified_name", ""))
        obj_rels = rels_by_obj.get(qn, [])

        if kind in ("table", "view"):
            out.append(f"\n-- {_ddl_comment_safe(qn)} ({kind})")
            ev = e.get("evidence") or {}
            if ev.get("file_path"):
                out.append(f"--   source: {_ddl_comment_safe(ev.get('file_path'))}:{ev.get('line', '')}")

            col_lines: list[str] = []
            for f in _sorted_fields(e):
                # Only emit a live column when the field is grounded (declared).
                if f.get("confidence") != "grounded":
                    col_lines.append(
                        f"  -- INFERRED COLUMN ({f.get('confidence','')}): "
                        f"{_ddl_comment_safe(f.get('name',''))} {_ddl_comment_safe(_field_type_display(f))}"
                    )
                    continue
                typ = _field_type_display(f) or "/* type unknown */"
                nn = ""
                if f.get("nullable") is False and f.get("enforcement") == "declared":
                    nn = " NOT NULL"
                col_lines.append(f"  {_ddl_ident(f.get('name',''))} {typ}{nn},")

            # Grounded, declared PK / UNIQUE constraints (table-level)
            constraint_lines: list[str] = []
            for r in obj_rels:
                if r.get("kind") == "pk" and r.get("confidence") == "grounded" and r.get("enforcement") == "declared":
                    constraint_lines.append(f"  PRIMARY KEY ({_ddl_ident(r.get('from_field',''))}),")
                elif r.get("kind") == "unique" and r.get("confidence") == "grounded" and r.get("enforcement") == "declared":
                    constraint_lines.append(f"  UNIQUE ({_ddl_ident(r.get('from_field',''))}),")
                elif _is_grounded_declared_fk(r):
                    constraint_lines.append(
                        f"  FOREIGN KEY ({_ddl_ident(r.get('from_field',''))}) "
                        f"REFERENCES {_ddl_ident(r.get('to_object',''))} ({_ddl_ident(r.get('to_field',''))}),"
                    )

            body = col_lines + constraint_lines
            if body:
                # strip the trailing comma on the last *live* (non-comment) line
                last_live = None
                for i in range(len(body) - 1, -1, -1):
                    if not body[i].lstrip().startswith("--"):
                        last_live = i
                        break
                if last_live is not None and body[last_live].rstrip().endswith(","):
                    body[last_live] = body[last_live].rstrip()[:-1]
                out.append(f"CREATE TABLE {_ddl_ident(qn)} (")
                out.extend(body)
                out.append(");")
            else:
                out.append(f"-- CREATE TABLE {_ddl_comment_safe(qn)} (no grounded columns — see HTML/CSV)")

            # Inferred/speculative relationships for this object -> commented
            for r in obj_rels:
                if r.get("kind") in ("fk", "join") and not _is_grounded_declared_fk(r):
                    out.append(_inferred_fk_comment(qn, r))

        else:
            # cobol_record / flatfile_layout: commented field manifest, not a table.
            out.append(f"\n-- {_ddl_comment_safe(qn)} ({kind}) — not relational; field manifest only")
            if e.get("variable_length"):
                out.append("--   record length: VARIABLE (ODO)")
            elif e.get("record_length") is not None:
                out.append(f"--   record length: {e.get('record_length')} bytes")
            for f in _sorted_fields(e):
                out.append(
                    f"--   {_ddl_comment_safe(f.get('name',''))} "
                    f"off={_ddl_comment_safe(_offset_display(f))} "
                    f"len={_ddl_comment_safe(_length_display(f))} "
                    f"[{f.get('confidence','')}]"
                )
            # Advisory relationships OWNED by a non-relational entity (e.g. a
            # speculative COBOL cross-record FK) must STILL be surfaced — never
            # silently dropped. They are commented-only, never live.
            for r in obj_rels:
                if r.get("kind") in ("fk", "join") and not _is_grounded_declared_fk(r):
                    out.append(_inferred_fk_comment(qn, r))

    out.append("")  # trailing newline
    ddl_path = output_dir / "structure.ddl.sql"
    atomic_write_text(ddl_path, "\n".join(out))
    return ddl_path


# ---------------------------------------------------------------------------
# Excel (.xlsx) — WP-8, COMPOSES ms-office-excel-python (openpyxl + safe_cell)
# ---------------------------------------------------------------------------

def safe_cell(value: Any) -> Any:
    """Formula-injection (CWE-1236) sanitization — INHERITED VERBATIM from the
    ms-office-excel-python skill HARD-RULE 1. A string cell whose first character
    is a formula trigger (=, +, -, @, TAB, CR) is prefixed with a single
    apostrophe, which forces Excel to interpret the cell as text. Non-string
    values pass through unchanged.

    This is BROADER than the CSV neutraliser ``_csv_escape`` (which only guards
    =,+,-,@): Excel additionally treats a leading TAB / CR as a formula lead-in,
    so the xlsx path must guard \\t and \\r too. We do NOT reinvent the rule — it
    is the skill's exact predicate.
    """
    if isinstance(value, str) and value and value[0] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + value  # leading apostrophe forces text interpretation
    return value


# openpyxl is an OPTIONAL dependency. The xlsx emitter degrades gracefully when
# it is absent (the catalog still has HTML/CSV/DDL/wiki). The safe_cell rule
# above is import-free so it is unit-testable without openpyxl.
try:  # pragma: no cover - import availability is environment-dependent
    import openpyxl as _openpyxl  # noqa: F401
    _OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover
    _openpyxl = None
    _OPENPYXL_AVAILABLE = False


def openpyxl_available() -> bool:
    """True iff openpyxl can be imported in this environment (test/skip guard)."""
    return _OPENPYXL_AVAILABLE


def _excel_sheet_title(qn: str, used: set[str]) -> str:
    """A valid, unique openpyxl sheet title.

    Excel sheet-name rules: <= 31 chars, none of ``: \\ / ? * [ ]``, not blank.
    On collision (e.g. two long qualified names sharing a 31-char prefix) we
    append a numeric suffix. Deterministic given the iteration order (entities
    are pre-sorted by the caller).
    """
    cleaned = "".join("_" if c in ':\\/?*[]' else c for c in str(qn)) or "entity"
    base = cleaned[:31]
    title = base
    n = 1
    while title.lower() in used:
        suffix = f"_{n}"
        title = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(title.lower())
    return title


def render_excel(catalog: dict, output_dir: Path, *, filename: str = "structure.xlsx") -> Optional[Path]:
    """Write structure.xlsx: one sheet per entity + a Summary sheet + a
    Relationships sheet. COMPOSES ms-office-excel-python (openpyxl) and applies
    the inherited ``safe_cell`` sanitization to EVERY user-controlled string
    cell (entity names, field names, types, evidence) — never writes a raw
    leading =,+,-,@,TAB,CR.

    Returns the output path, or ``None`` when openpyxl is unavailable (the
    caller treats this as a sanctioned skip — the rest of the catalog is intact).
    Document-property leakage is avoided by stripping creator / lastModifiedBy
    (ms-office-excel-python Security Hardening: workbook properties carry PII).
    """
    if not _OPENPYXL_AVAILABLE:
        return None

    from openpyxl import Workbook  # local import — only when actually emitting

    output_dir.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # Strip identity-bearing workbook properties before anything is written
    # (Security Hardening item: creator/last_modified_by carry user-identifiable
    # data — sanitize before publishing externally). We pin created/modified to a
    # FIXED epoch rather than clearing them: openpyxl's core-properties serializer
    # requires datetimes (a None breaks the W3CDTF timestamp element), and a fixed
    # value keeps the xlsx free of real wall-clock identity. Deterministic output.
    try:
        _fixed = _dt.datetime(2020, 1, 1, 0, 0, 0, tzinfo=_dt.timezone.utc)
        wb.properties.creator = "structure-recovery"
        wb.properties.lastModifiedBy = "structure-recovery"
        wb.properties.created = _fixed
        wb.properties.modified = _fixed
    except Exception:
        pass

    entities = _sorted_entities(catalog)
    relationships = _sorted_relationships(catalog)

    # --- Summary sheet (first; reuse the default sheet) ---
    summary = wb.active
    summary.title = "Summary"
    summary.append(["entity", "object_kind", "confidence", "field_count", "source"])
    for e in entities:
        ev = e.get("evidence") or {}
        src = ""
        if ev.get("file_path"):
            ln = ev.get("line")
            src = f"{ev.get('file_path')}:{ln}" if ln is not None else str(ev.get("file_path"))
        summary.append([
            safe_cell(e.get("qualified_name", "")),
            safe_cell(e.get("object_kind", "")),
            safe_cell(e.get("confidence", "")),
            len(_sorted_fields(e)),  # int — no sanitization needed
            safe_cell(src),
        ])

    # --- One sheet per entity ---
    used_titles: set[str] = {"summary", "relationships"}
    for e in entities:
        title = _excel_sheet_title(str(e.get("qualified_name", "")), used_titles)
        ws = wb.create_sheet(title=title)
        has_cobol = e.get("object_kind") == "cobol_record"
        if has_cobol:
            ws.append(["field", "level", "type", "offset", "length", "confidence", "evidence_kind"])
        else:
            ws.append(["column", "type", "nullable", "confidence", "evidence_kind", "enforcement"])
        for f in _sorted_fields(e):
            if has_cobol:
                level = f.get("level")
                ws.append([
                    safe_cell(f.get("name", "")),
                    level if level is not None else "",  # int/blank
                    safe_cell(_field_type_display(f)),
                    safe_cell(_offset_display(f)),  # str 'min..max' or '' — str path, sanitized
                    safe_cell(_length_display(f)),
                    safe_cell(f.get("confidence", "")),
                    safe_cell(f.get("evidence_kind", "")),
                ])
            else:
                nullable = f.get("nullable")
                null_disp = "" if nullable is None else ("YES" if nullable else "NO")
                ws.append([
                    safe_cell(f.get("name", "")),
                    safe_cell(_field_type_display(f)),
                    safe_cell(null_disp),
                    safe_cell(f.get("confidence", "")),
                    safe_cell(f.get("evidence_kind", "")),
                    safe_cell(f.get("enforcement", "")),
                ])

    # --- Relationships sheet ---
    rels_ws = wb.create_sheet(title="Relationships")
    rels_ws.append(["kind", "from_object", "from_field", "to_object", "to_field", "confidence", "enforcement"])
    for r in relationships:
        rels_ws.append([
            safe_cell(r.get("kind", "")),
            safe_cell(r.get("from_object", "")),
            safe_cell(r.get("from_field", "")),
            safe_cell(r.get("to_object") or ""),
            safe_cell(r.get("to_field") or ""),
            safe_cell(r.get("confidence", "")),
            safe_cell(r.get("enforcement", "")),
        ])

    # Atomic save: write to a sibling .tmp then os.replace (mirrors
    # atomic_write_text). openpyxl.save wants a path; we give it the temp path.
    out_path = output_dir / filename
    fd, tmp_path = tempfile.mkstemp(
        prefix=out_path.name + ".tmp.", suffix=f".{os.getpid()}", dir=str(output_dir),
    )
    os.close(fd)  # openpyxl re-opens the path itself
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, out_path)
    except Exception:
        try:
            os.unlink(tmp_path)
        except FileNotFoundError:
            pass
        raise
    return out_path


# ---------------------------------------------------------------------------
# Wiki (markdown pages) — WP-8, COMPOSES the wiki skill
#   (per-page file:line citations + .wiki.lock + anti-pollution default)
# ---------------------------------------------------------------------------

class WikiAntiPollutionError(RuntimeError):
    """Raised when render_wiki is pointed at a SHARED wiki without an explicit
    opt-in. The wiki skill's anti-pollution default forbids auto-filing NEW
    pages into a role: shared wiki without user approval."""


def _wiki_page_slug(qn: str) -> str:
    """A filesystem-safe, deterministic page slug for an entity."""
    s = "".join(c if (c.isalnum() or c in ("-", "_", ".")) else "-" for c in str(qn))
    s = s.strip("-") or "entity"
    return s[:120]


def _wiki_cite(evidence: Optional[dict]) -> str:
    """Render a `[Source: file:line]` citation from an evidence block, honoring
    the wiki HARD-RULE 'cite every claim'. file:line is the structure-recovery
    equivalent of the wiki's `raw/<file>, p.<page>` form. Returns '' when no
    evidence is available (the fact is then explicitly marked uncited)."""
    if not evidence:
        return ""
    fp = evidence.get("file_path")
    if not fp:
        return ""
    ln = evidence.get("line")
    loc = f"{fp}:{ln}" if ln is not None else str(fp)
    return f"[Source: {loc}]"


def _wiki_acquire_lock(wiki_root: Path) -> bool:
    """Acquire the single-writer `.wiki.lock` (wiki HARD-RULE). Atomic create via
    O_CREAT|O_EXCL — returns False if another writer holds it. Reads never call
    this."""
    lock_path = wiki_root / ".wiki.lock"
    try:
        fd = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY, 0o644)
    except FileExistsError:
        return False
    try:
        os.write(fd, f"structure-recovery pid={os.getpid()}\n".encode("utf-8"))
    finally:
        os.close(fd)
    return True


def _wiki_release_lock(wiki_root: Path) -> None:
    """Release the `.wiki.lock` (best-effort; called in finally)."""
    try:
        (wiki_root / ".wiki.lock").unlink()
    except FileNotFoundError:
        pass


def render_wiki(
    catalog: dict,
    output_dir: Path,
    *,
    project_name: str = "project",
    wiki_role: str = "specific",
    allow_shared_write: bool = False,
    source_date_epoch: Optional[int] = None,
) -> dict[str, Any]:
    """Write one interlinked markdown wiki page per entity + an index page, with
    EVERY structural fact cited to file:line. COMPOSES the wiki skill:

      * cite-every-claim    — each entity/field fact carries `[Source: file:line]`.
      * single-writer lock  — pages are written under `.wiki.lock` (acquire,
                              write, release in finally). Reads never lock.
      * anti-pollution      — a SHARED wiki (role != 'specific') is NOT written
                              without an explicit ``allow_shared_write=True``
                              opt-in; otherwise WikiAntiPollutionError is raised
                              BEFORE the lock is acquired. This honors the wiki
                              default that NEW pages into a shared wiki need
                              user approval.

    The pages are the structure-recovery deliverable surface (a self-contained
    `wiki/` page set under ``output_dir``), NOT an auto-ingest into the user's
    personal knowledge wiki — so the anti-pollution guard is a real boundary,
    not a formality.

    Returns {'pages': {qn: path}, 'index': path, 'wiki_dir': dir, 'locked': bool}.
    """
    # Anti-pollution guard FIRST — refuse before touching the filesystem.
    if wiki_role != "specific" and not allow_shared_write:
        raise WikiAntiPollutionError(
            f"refusing to auto-file structure pages into a role={wiki_role!r} wiki "
            f"without allow_shared_write=True (wiki anti-pollution default)"
        )

    sde = _resolve_source_date_epoch(source_date_epoch)
    wiki_dir = output_dir / "wiki"
    wiki_dir.mkdir(parents=True, exist_ok=True)

    entities = _sorted_entities(catalog)
    relationships = _sorted_relationships(catalog)
    rels_by_obj: dict[str, list[dict]] = {}
    for r in relationships:
        rels_by_obj.setdefault(str(r.get("from_object", "")), []).append(r)

    # Pre-compute slugs so pages can interlink each other deterministically.
    slugs: dict[str, str] = {}
    used_slugs: set[str] = set()
    for e in entities:
        qn = str(e.get("qualified_name", ""))
        slug = _wiki_page_slug(qn)
        base = slug
        n = 1
        while slug in used_slugs:
            slug = f"{base}-{n}"
            n += 1
        used_slugs.add(slug)
        slugs[qn] = slug

    locked = _wiki_acquire_lock(wiki_dir)
    produced: dict[str, Path] = {}
    index_path = wiki_dir / "index.md"
    try:
        # --- per-entity pages ---
        for e in entities:
            qn = str(e.get("qualified_name", ""))
            kind = str(e.get("object_kind", ""))
            ev = e.get("evidence") or {}
            cite = _wiki_cite(ev)
            lines: list[str] = []
            lines.append(f"# {qn}")
            lines.append("")
            lines.append(f"- **kind**: `{kind}`")
            conf_cite = f" {cite}" if cite else "  _(uncited — no source evidence)_"
            lines.append(f"- **confidence**: `{e.get('confidence', '')}`{conf_cite}")
            if e.get("variable_length"):
                lo = e.get("record_length_min")
                hi = e.get("record_length_max")
                rng = f"{lo if lo is not None else '?'}..{hi if hi is not None else '?'}"
                lines.append(f"- **record length**: varies (ODO) {rng} bytes")
            elif e.get("record_length") is not None:
                lines.append(f"- **record length**: {e.get('record_length')} bytes")
            lines.append("")
            lines.append("> Inferred structure — characterization aid; requires human review.")
            lines.append("")

            # Fields table with a per-field citation column.
            has_cobol = kind == "cobol_record"
            if has_cobol:
                lines.append("## Fields")
                lines.append("")
                lines.append("| field | level | type | offset | length | confidence | source |")
                lines.append("|---|---|---|---|---|---|---|")
            else:
                lines.append("## Columns")
                lines.append("")
                lines.append("| column | type | nullable | confidence | enforcement | source |")
                lines.append("|---|---|---|---|---|---|")
            for f in _sorted_fields(e):
                fcite = _wiki_cite(f.get("evidence")) or cite or "_(uncited)_"
                fcite = _md_cell(fcite)
                if has_cobol:
                    level = f.get("level")
                    lines.append(
                        f"| {_md_cell(f.get('name', ''))} "
                        f"| {_md_cell('' if level is None else level)} "
                        f"| {_md_cell(_field_type_display(f))} "
                        f"| {_md_cell(_offset_display(f))} "
                        f"| {_md_cell(_length_display(f))} "
                        f"| {_md_cell(f.get('confidence', ''))} "
                        f"| {fcite} |"
                    )
                else:
                    nullable = f.get("nullable")
                    null_disp = "" if nullable is None else ("YES" if nullable else "NO")
                    lines.append(
                        f"| {_md_cell(f.get('name', ''))} "
                        f"| {_md_cell(_field_type_display(f))} "
                        f"| {_md_cell(null_disp)} "
                        f"| {_md_cell(f.get('confidence', ''))} "
                        f"| {_md_cell(f.get('enforcement', ''))} "
                        f"| {fcite} |"
                    )
            lines.append("")

            # Relationships OWNED by this entity, interlinked to target pages.
            obj_rels = sorted(rels_by_obj.get(qn, []), key=_rel_sort_key)
            if obj_rels:
                lines.append("## Relationships")
                lines.append("")
                for r in obj_rels:
                    tgt_qn = r.get("to_object")
                    if tgt_qn and str(tgt_qn) in slugs:
                        tgt = f"[{tgt_qn}]({slugs[str(tgt_qn)]}.md)"  # interlink
                    elif tgt_qn:
                        tgt = f"`{tgt_qn}`"  # target not a catalog entity — no link
                    else:
                        tgt = "_(unresolved)_"
                    rcite = _wiki_cite(r.get("evidence"))
                    rcite_s = f" {rcite}" if rcite else ""
                    advisory = "" if (
                        r.get("kind") == "fk" and r.get("confidence") == "grounded"
                        and r.get("enforcement") == "declared"
                    ) else "  _(advisory)_"
                    lines.append(
                        f"- `{r.get('kind', '')}` "
                        f"`{r.get('from_field', '')}` → {tgt} "
                        f"`{r.get('to_field') or '?'}` "
                        f"[{r.get('confidence', '')}]{rcite_s}{advisory}"
                    )
                lines.append("")

            # Per-entity gaps (also cited where a location is known).
            egaps = e.get("gaps") or []
            if egaps:
                lines.append("## Gaps")
                lines.append("")
                for g in sorted(egaps, key=lambda x: (str(x.get("kind", "")),
                                                       str(x.get("file_path") or ""),
                                                       x.get("line") or 0)):
                    gcite = _wiki_cite({"file_path": g.get("file_path"), "line": g.get("line")})
                    gcite_s = f" {gcite}" if gcite else ""
                    lines.append(f"- `{g.get('kind', '')}` — {_md_inline(g.get('description', ''))}{gcite_s}")
                lines.append("")

            lines.append(f"---")
            lines.append(f"_[index](index.md) · generated by {EXTRACTOR_ID} v{EXTRACTOR_VERSION} at {_generated_at(sde)}_")
            lines.append("")
            page_path = wiki_dir / f"{slugs[qn]}.md"
            atomic_write_text(page_path, "\n".join(lines))
            produced[qn] = page_path

        # --- index page (interlinks every entity page) ---
        idx: list[str] = []
        idx.append(f"# Structure catalog — {project_name}")
        idx.append("")
        idx.append(f"_{len(entities)} entit{'y' if len(entities) == 1 else 'ies'}, "
                   f"{len(relationships)} relationship{'' if len(relationships) == 1 else 's'}. "
                   f"Generated by {EXTRACTOR_ID} v{EXTRACTOR_VERSION} at {_generated_at(sde)}._")
        idx.append("")
        idx.append("> Inferred structure — characterization aid; every fact is cited to file:line. "
                   "Inferred/speculative relationships are advisory only.")
        idx.append("")
        idx.append("## Entities")
        idx.append("")
        for e in entities:
            qn = str(e.get("qualified_name", ""))
            ev = e.get("evidence") or {}
            cite = _wiki_cite(ev)
            cite_s = f" — {cite}" if cite else ""
            idx.append(f"- [{qn}]({slugs[qn]}.md) "
                       f"`{e.get('object_kind', '')}` "
                       f"[{e.get('confidence', '')}]{cite_s}")
        idx.append("")
        atomic_write_text(index_path, "\n".join(idx))
    finally:
        if locked:
            _wiki_release_lock(wiki_dir)

    return {"pages": produced, "index": index_path, "wiki_dir": wiki_dir, "locked": locked}


# ---------------------------------------------------------------------------
# WP-9 — OpenLineage SchemaDatasetFacet emission (M1)
#
# COMPOSES lineage-extract-static/scripts/merge_into_ol.py (does NOT reinvent OL
# emission). The vendored OL 2.0.2 schema accepts a `schema` facet on a
# DatasetEvent (DatasetRef.facets is an open object) — verified at design time
# and re-asserted in test_ol_schema_facet.py. Each structure-index entity is
# projected into a SchemaDatasetFacet (fields + types) that ENRICHES the EXISTING
# lineage DatasetEvent for the same dataset. Fail-closed: if the enriched event
# would be rejected by the vendored schema, the facet is dropped, the bare event
# is kept, and a gap is recorded — never a malformed emit, never a whole-event
# abort (design §9 note 3).
# ---------------------------------------------------------------------------

# Lazily import the sibling merge_into_ol (it lives in lineage-extract-static).
# Import-by-path keeps render_structure.py runnable from the repo OR the shadow
# (~/.claude) tree, and keeps OL an OPTIONAL compose seam — when the sibling /
# jsonschema is unavailable, render_ol_schema_facets degrades gracefully.
def _import_merge_into_ol():
    """Return the merge_into_ol module, or None if it (or jsonschema) is absent.

    The lineage sibling sits at ../../lineage-extract-static/scripts/ relative to
    this file in BOTH the repo and the ~/.claude shadow layout.
    """
    import importlib.util as _ilu

    here = Path(__file__).resolve()
    sibling = (
        here.parent.parent.parent
        / "lineage-extract-static"
        / "scripts"
        / "merge_into_ol.py"
    )
    if not sibling.exists():
        return None
    try:
        # merge_into_ol does `sys.path.insert(0, <its own dir>)` then
        # `from validate_ol import ...` at module load, so its sibling validator
        # resolves regardless of our sys.path.
        spec = _ilu.spec_from_file_location("merge_into_ol", sibling)
        if not spec or not spec.loader:
            return None
        mod = _ilu.module_from_spec(spec)
        sys.modules.setdefault("merge_into_ol", mod)
        spec.loader.exec_module(mod)
        return mod
    except Exception:
        # jsonschema missing, validator import failure, etc. — OL facet is optional.
        return None


def _entity_dataset_name(entity: dict) -> str:
    """Dataset identity used to match a structure entity to a lineage dataset.

    We key on ``qualified_name`` — the canonical entity identity in
    structure-index.v1 — so the facet enriches the lineage DatasetEvent whose
    ``dataset.name`` equals the entity's qualified name. Deterministic, no
    fabrication of namespace.
    """
    return str(entity.get("qualified_name", ""))


def _facet_fields_for_entity(entity: dict) -> list[dict]:
    """Project an entity's real columns into OL SchemaDatasetFacet ``fields``.

    Skips structural scaffolding that is not an emittable column: COBOL group
    nodes (``is_group``) and FILLER (``is_filler``) carry no data type of their
    own. Each emitted field is ``{name, type[, description]}`` where ``type`` is
    the best declared/pic/normalized label via the existing ``_field_type_display``
    helper (REUSED — no second type-precedence rule). Order follows the catalog's
    deterministic field sort.
    """
    out: list[dict] = []
    for f in _sorted_fields(entity):
        if f.get("is_group") or f.get("is_filler"):
            continue
        name = str(f.get("name", ""))
        if not name:
            continue
        fld: dict = {"name": name, "type": _field_type_display(f) or "unknown"}
        # Carry the COBOL PIC as a description when it adds info beyond `type`.
        pic = f.get("pic_clause")
        if pic and str(pic) != fld["type"]:
            fld["description"] = f"PIC {pic}"
        out.append(fld)
    return out


def build_schema_facets_from_catalog(catalog: dict) -> "dict[str, dict]":
    """Build ``{dataset_name: SchemaDatasetFacet}`` for every catalog entity.

    Composes ``merge_into_ol.make_schema_dataset_facet``. Returns an EMPTY dict
    (not an error) when the sibling OL module is unavailable — the caller treats
    that as a sanctioned skip. Entities with zero emittable fields are omitted
    (an empty schema facet enriches nothing).
    """
    mio = _import_merge_into_ol()
    if mio is None:
        return {}
    facets: dict[str, dict] = {}
    for e in _sorted_entities(catalog):
        name = _entity_dataset_name(e)
        if not name:
            continue
        fields = _facet_fields_for_entity(e)
        if not fields:
            continue
        facets[name] = mio.make_schema_dataset_facet(fields)
    return facets


def render_ol_schema_facets(
    catalog: dict,
    output_dir: Path,
    *,
    filename: str = "structure.schema-facets.json",
) -> Optional[Path]:
    """Emit the structure SchemaDatasetFacets as a JSON artifact AND a parallel
    set of enriched, schema-validated DatasetEvents (M1).

    The artifact is ``{"schema_facets": {dataset_name: facet, ...}, "datasets":
    [enriched-or-bare DatasetEvent, ...], "gaps": [...]}``. Every DatasetEvent is
    passed through ``attach_schema_facet_fail_closed`` so the emitted facet is
    GUARANTEED to validate against the vendored OL 2.0.2 schema; a rejected facet
    is dropped (bare event kept) with a recorded gap — never a malformed emit,
    never a whole-event abort (design §9 note 3 / AC3).

    Returns the artifact path, or ``None`` when the OL sibling is unavailable
    (sanctioned skip — the rest of the catalog is intact). Deterministic, atomic.
    """
    mio = _import_merge_into_ol()
    if mio is None:
        return None

    facets = build_schema_facets_from_catalog(catalog)

    # Fixed eventTime for byte-deterministic output (this is a static, runtime-free
    # facet emission — wall-clock identity is intentionally stripped, mirroring the
    # WP-7 SOURCE_DATE_EPOCH discipline and the WP-8 fixed-epoch workbook props).
    scan_started_at = "2020-01-01T00:00:00.000Z"

    enriched_events: list[dict] = []
    gaps: list[dict] = []
    # Deterministic order: sort by dataset name.
    for e in _sorted_entities(catalog):
        name = _entity_dataset_name(e)
        if not name:
            continue
        # The lineage dataset namespace is not carried by structure-index; use a
        # stable structure-recovery namespace so the enriched DatasetEvent is
        # self-consistent and schema-valid. (When fed alongside a real lineage
        # run, consumers match on dataset.name — the facet is the payload.)
        dataset_ref = {
            "namespace": "structure://" + str(catalog.get("extractor_id", "structure-recovery")),
            "name": name,
            "kind": "table" if e.get("object_kind") in ("table", "view") else "file",
        }
        facet = facets.get(name)
        # Build the bare event first (always schema-valid), then fail-closed attach.
        bare = mio.make_dataset_event(dataset_ref, scan_started_at)
        if facet is None:
            enriched_events.append(bare)
            continue
        event, gap_reason = mio.attach_schema_facet_fail_closed(bare, facet)
        enriched_events.append(event)
        if gap_reason is not None:
            gaps.append({"dataset": name, "kind": "schema_facet_rejected", "detail": gap_reason})

    artifact = {
        "schema_facets": facets,
        "datasets": enriched_events,
        "gaps": gaps,
    }
    out_path = output_dir / filename
    atomic_write_text(
        out_path,
        json.dumps(artifact, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
    )
    return out_path


# ---------------------------------------------------------------------------
# Orchestration + CLI
# ---------------------------------------------------------------------------

def render_all(
    catalog: dict,
    output_dir: Path,
    *,
    no_vendor: bool = False,
    project_name: str = "project",
    source_date_epoch: Optional[int] = None,
) -> dict[str, Any]:
    """Render the three stdlib outputs. Returns a dict of produced paths.

    WP-8 (Excel/wiki) and WP-9 (OL facet) compose render_csv / render_ddl /
    this catalog separately — they are NOT invoked here.
    """
    html_path = render_html(
        catalog, output_dir, no_vendor=no_vendor,
        project_name=project_name, source_date_epoch=source_date_epoch,
    )
    csv_paths = render_csv(catalog, output_dir)
    ddl_path = render_ddl(catalog, output_dir)
    return {
        "html": html_path,
        "fields_csv": csv_paths["fields.csv"],
        "relationships_csv": csv_paths["relationships.csv"],
        "ddl": ddl_path,
    }


def main(argv: Optional[list[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("catalog_path", type=Path, help="Path to structure-index.v1 catalog JSON")
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--no-vendor", action="store_true", help="Do not fall back to CDN; degrade graph to a banner if vendor absent")
    parser.add_argument("--project-name", default="project")
    parser.add_argument("--source-date-epoch", type=int, default=None)
    args = parser.parse_args(argv)

    if not args.catalog_path.exists():
        print(f"ERROR: catalog not found: {args.catalog_path}", file=sys.stderr)
        return 1
    try:
        with args.catalog_path.open("r", encoding="utf-8") as f:
            catalog = json.load(f)
    except json.JSONDecodeError as e:
        print(f"ERROR: failed to parse catalog: {e}", file=sys.stderr)
        return 1

    try:
        produced = render_all(
            catalog,
            args.output_dir,
            no_vendor=args.no_vendor,
            project_name=args.project_name,
            source_date_epoch=args.source_date_epoch,
        )
    except (OSError, PermissionError) as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 1

    print(json.dumps({k: str(v) for k, v in produced.items()}, sort_keys=True))
    return 0


if __name__ == "__main__":
    sys.exit(main())
